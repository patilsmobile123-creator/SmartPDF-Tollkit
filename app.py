from flask import Flask, request, send_file, jsonify, g
from flask_cors import CORS
import os
import io
import zipfile
import tempfile
import time
import csv
import json
from datetime import datetime
from functools import wraps
import threading
from pypdf import PdfReader, PdfWriter
from PIL import Image
from reportlab.lib.pagesizes import A4, letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import fitz  # PyMuPDF
import requests
import hashlib


app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

# Admin password — set ADMIN_PASSWORD env var on Render (never hardcode!)
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', '')
if not ADMIN_PASSWORD:
    import warnings
    warnings.warn("ADMIN_PASSWORD env var is not set — all /admin/* endpoints are disabled.")

# Paths
UPLOAD_FOLDER = tempfile.mkdtemp()
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# CSV file for visitor logs — use /tmp/ so it works on Render and similar hosts
VISITOR_LOG_FILE = os.path.join(tempfile.gettempdir(), 'visitor_logs.csv')

# Ensure CSV exists with headers
if not os.path.exists(VISITOR_LOG_FILE):
    with open(VISITOR_LOG_FILE, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            'timestamp', 'date', 'time', 'ip_address', 'anonymous_id', 
            'tool_used', 'file_count', 'file_names', 'file_sizes_mb',
            'user_agent', 'referrer', 'country', 'city', 'page_url',
            'processing_time_ms', 'success', 'error_message'
        ])

# ── Google Sheets persistent logging ─────────────────────────────────────────
# Set these 2 env vars on Render to enable (see Step 5 setup guide):
#   SHEETS_CREDENTIALS  — contents of your service account JSON key (single line)
#   SHEETS_SPREADSHEET_ID — the ID from your Google Sheet URL
SHEETS_ENABLED = False
_sheets_service = None

def _init_sheets():
    """Lazy-init the Sheets client once on first use."""
    global SHEETS_ENABLED, _sheets_service
    creds_json = os.environ.get('SHEETS_CREDENTIALS', '')
    spreadsheet_id = os.environ.get('SHEETS_SPREADSHEET_ID', '')
    if not creds_json or not spreadsheet_id:
        return  # env vars not set — silently skip
    try:
        import json as _json
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
        creds_data = _json.loads(creds_json)
        creds = Credentials.from_service_account_info(
            creds_data,
            scopes=['https://www.googleapis.com/auth/spreadsheets']
        )
        _sheets_service = build('sheets', 'v4', credentials=creds, cache_discovery=False)
        SHEETS_ENABLED = True
        print("Google Sheets logging enabled.")
    except Exception as e:
        print(f"Google Sheets init failed: {e}")

# Try to init on startup
threading.Thread(target=_init_sheets, daemon=True).start()

def _append_to_sheet(row):
    """Append one row to Google Sheets — always runs in a background thread."""
    if not SHEETS_ENABLED or _sheets_service is None:
        return
    try:
        spreadsheet_id = os.environ.get('SHEETS_SPREADSHEET_ID', '')
        _sheets_service.spreadsheets().values().append(
            spreadsheetId=spreadsheet_id,
            range='Sheet1!A1',
            valueInputOption='RAW',
            insertDataOption='INSERT_ROWS',
            body={'values': [[str(v) for v in row]]}
        ).execute()
    except Exception as e:
        print(f"Sheets append error: {e}")

# Simple in-memory rate limiting (resets on restart)
request_times = {}

def log_visitor(tool_name='', file_count=0, file_names='', file_sizes=0, 
                success=True, error_msg='', processing_time=0):
    """Log visitor activity to CSV"""
    try:
        now = datetime.now()
        
        # Get IP (handle proxies)
        ip = request.headers.get('X-Forwarded-For', request.remote_addr)
        if ip and ',' in ip:
            ip = ip.split(',')[0].strip()
        
        # Generate anonymous visitor ID (hashed IP + user agent)
        visitor_hash = hashlib.sha256(f"{ip}{request.user_agent.string}".encode()).hexdigest()[:16]
        
        # Write row immediately — background thread fills in country/city later
        row_timestamp = now.isoformat()
        row = [
            row_timestamp,
            now.strftime('%Y-%m-%d'),
            now.strftime('%H:%M:%S'),
            ip or 'unknown',
            visitor_hash,
            tool_name,
            file_count,
            file_names[:500],
            round(file_sizes, 2),
            request.user_agent.string[:200] if request.user_agent else 'unknown',
            request.referrer or 'direct',
            'unknown',  # country — filled by background thread
            'unknown',  # city   — filled by background thread
            request.url,
            processing_time,
            'yes' if success else 'no',
            error_msg[:200] if error_msg else ''
        ]
        
        with open(VISITOR_LOG_FILE, 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(row)

        # Fire geolocation in background — does NOT block the user response
        t = threading.Thread(target=_fetch_and_update_location, args=(ip, row_timestamp), daemon=True)
        t.start()

        # Also append to Google Sheets in background (permanent storage)
        t2 = threading.Thread(target=_append_to_sheet, args=(row,), daemon=True)
        t2.start()
            
    except Exception as e:
        print(f"Logging error: {e}")

def get_location_from_ip(ip):
    """Returns instantly — real location is filled in by background thread."""
    return 'unknown', 'unknown'

def _fetch_and_update_location(ip, row_timestamp):
    """Runs in a background thread — fetches IP location and updates the CSV row."""
    try:
        response = requests.get(f'https://ipinfo.io/{ip}/json', timeout=3)
        if response.status_code == 200:
            data = response.json()
            country = data.get('country', 'unknown')
            city = data.get('city', 'unknown')
            if country == 'unknown' and city == 'unknown':
                return
            # Update the matching row in the CSV (country=col12, city=col13)
            rows = []
            with open(VISITOR_LOG_FILE, 'r', newline='', encoding='utf-8') as f:
                rows = list(csv.reader(f))
            for i, row in enumerate(rows):
                if i > 0 and len(row) > 13 and row[0] == row_timestamp:
                    rows[i][12] = country
                    rows[i][13] = city
                    break
            with open(VISITOR_LOG_FILE, 'w', newline='', encoding='utf-8') as f:
                csv.writer(f).writerows(rows)
    except Exception as e:
        print(f"Background location fetch error: {e}")

def track_usage(tool_name):
    """Decorator to track tool usage"""
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            start_time = time.time()
            success = True
            error_msg = ''
            file_count = 0
            file_names = ''
            total_size = 0
            
            try:
                # Count files before processing
                if 'files' in request.files:
                    files = request.files.getlist('files') or request.files.getlist('pdfs') or request.files.getlist('images')
                    file_count = len(files)
                    file_names = ', '.join([f.filename for f in files if f])
                    total_size = sum([len(f.read()) for f in files if f]) / (1024*1024)  # MB
                    # Reset file pointers
                    for f in files:
                        f.seek(0)
                elif request.files:
                    file_count = len(request.files)
                    files = list(request.files.values())
                    file_names = ', '.join([f.filename for f in files if f])
                    total_size = sum([len(f.read()) for f in files if f]) / (1024*1024)
                    for f in files:
                        f.seek(0)
                        
                result = f(*args, **kwargs)
                return result
                
            except Exception as e:
                success = False
                error_msg = str(e)
                raise e
            finally:
                processing_time = int((time.time() - start_time) * 1000)
                log_visitor(
                    tool_name=tool_name,
                    file_count=file_count,
                    file_names=file_names,
                    file_sizes=total_size,
                    success=success,
                    error_msg=error_msg,
                    processing_time=processing_time
                )
        return wrapper
    return decorator

# ADMIN ENDPOINT - Download CSV logs (add password protection!)
@app.route('/admin/logs')
def download_logs():
    """Download visitor logs as CSV (PROTECT THIS IN PRODUCTION!)"""
    password = request.args.get('password', '')
    
    # Simple password protection - CHANGE THIS!
    if not ADMIN_PASSWORD or password != ADMIN_PASSWORD:
        return jsonify({'error': 'Unauthorized'}), 401
    
    if not os.path.exists(VISITOR_LOG_FILE):
        return jsonify({'error': 'No logs found'}), 404
    
    return send_file(
        VISITOR_LOG_FILE,
        as_attachment=True,
        download_name=f'visitor_logs_{datetime.now().strftime("%Y%m%d")}.csv',
        mimetype='text/csv'
    )

# ADMIN ENDPOINT - View stats JSON
@app.route('/admin/stats')
def view_stats():
    """View usage statistics (PROTECT THIS!)"""
    password = request.args.get('password', '')
    if not ADMIN_PASSWORD or password != ADMIN_PASSWORD:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        stats = {
            'total_visits': 0,
            'unique_visitors': set(),
            'tool_usage': {},
            'daily_stats': {},
            'errors': 0
        }
        
        with open(VISITOR_LOG_FILE, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                stats['total_visits'] += 1
                stats['unique_visitors'].add(row['anonymous_id'])
                
                tool = row['tool_used'] or 'page_view'
                stats['tool_usage'][tool] = stats['tool_usage'].get(tool, 0) + 1
                
                date = row['date']
                if date not in stats['daily_stats']:
                    stats['daily_stats'][date] = {'visits': 0, 'unique': set()}
                stats['daily_stats'][date]['visits'] += 1
                stats['daily_stats'][date]['unique'].add(row['anonymous_id'])
                
                if row['success'] == 'no':
                    stats['errors'] += 1
        
        # Convert sets to counts for JSON serialization
        stats['unique_visitors'] = len(stats['unique_visitors'])
        for date in stats['daily_stats']:
            stats['daily_stats'][date]['unique'] = len(stats['daily_stats'][date]['unique'])
        
        return jsonify(stats)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ADMIN ENDPOINT - Download logs as Excel
@app.route('/admin/logs/excel')
def download_logs_excel():
    """Download visitor logs as a formatted Excel file"""
    password = request.args.get('password', '')
    if not ADMIN_PASSWORD or password != ADMIN_PASSWORD:
        return jsonify({'error': 'Unauthorized'}), 401

    if not os.path.exists(VISITOR_LOG_FILE):
        return jsonify({'error': 'No logs found'}), 404

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ── Sheet 1: Raw Logs ──────────────────────────────────────────
        ws = wb.active
        ws.title = 'Visitor Logs'

        rows = []
        with open(VISITOR_LOG_FILE, 'r', encoding='utf-8') as f:
            rows = list(csv.reader(f))

        if not rows:
            return jsonify({'error': 'No data in logs'}), 404

        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col_idx, cell_value in enumerate(rows[0], 1):
            cell = ws.cell(row=1, column=col_idx, value=cell_value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, row in enumerate(rows[1:], 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            # Alternating row color
            if row_idx % 2 == 0:
                for col_idx in range(1, len(row) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                        start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

        # Auto-size columns
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value or '')) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

        ws.freeze_panes = 'A2'

        # ── Sheet 2: Summary Stats ─────────────────────────────────────
        ws2 = wb.create_sheet('Summary')
        tool_counts = {}
        daily_counts = {}
        unique_visitors = set()
        total = 0
        errors = 0

        for row in rows[1:]:
            if len(row) < 17:
                continue
            total += 1
            unique_visitors.add(row[4])  # anonymous_id
            tool = row[5] or 'page_view'
            tool_counts[tool] = tool_counts.get(tool, 0) + 1
            date = row[1]
            daily_counts[date] = daily_counts.get(date, 0) + 1
            if row[16] == 'no':
                errors += 1

        # Summary header
        ws2['A1'] = 'SmartPDF Toolkit — Visitor Summary'
        ws2['A1'].font = Font(bold=True, size=14, color='1F4E79')
        ws2.merge_cells('A1:B1')

        summary_data = [
            ('', ''),
            ('Metric', 'Value'),
            ('Total Visits', total),
            ('Unique Visitors', len(unique_visitors)),
            ('Total Errors', errors),
            ('', ''),
            ('Tool', 'Uses'),
        ]
        for item in summary_data:
            ws2.append(item)

        # Style the "Metric/Value" and "Tool/Uses" header rows
        for r in ws2.iter_rows():
            if r[0].value in ('Metric', 'Tool'):
                for cell in r:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='2E86C1', end_color='2E86C1', fill_type='solid')

        for tool, count in sorted(tool_counts.items(), key=lambda x: -x[1]):
            ws2.append([tool, count])

        ws2.append(['', ''])
        ws2.append(['Date', 'Visits'])
        header_row = ws2.max_row
        for cell in ws2[header_row]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2E86C1', end_color='2E86C1', fill_type='solid')

        for date, count in sorted(daily_counts.items()):
            ws2.append([date, count])

        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 15

        # ── Save & send ───────────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'visitor_logs_{datetime.now().strftime("%Y%m%d")}.xlsx'
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except ImportError:
        return jsonify({'error': 'openpyxl not installed. Add it to requirements.txt'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Track page views
@app.route('/')
def home():
    log_visitor(tool_name='page_view')
    return app.send_static_file('index.html')

# 1. IMAGES TO PDF
@app.route('/api/images-to-pdf', methods=['POST'])
@track_usage('images_to_pdf')
def images_to_pdf():
    try:
        files = request.files.getlist('images')
        if not files:
            return jsonify({'error': 'No images uploaded'}), 400
        
        page_size = request.form.get('page_size', 'A4')
        orientation = request.form.get('orientation', 'portrait')
        fit_mode = request.form.get('fit_mode', 'fit')
        
        if page_size == 'A4':
            size = A4
        else:
            size = letter
            
        if orientation == 'landscape':
            size = (size[1], size[0])
        
        output_path = os.path.join(UPLOAD_FOLDER, f"images_{int(time.time())}.pdf")
        c = canvas.Canvas(output_path, pagesize=size)
        width, height = size
        
        for file in files:
            try:
                img = Image.open(file.stream)
                img_width, img_height = img.size
                aspect = img_height / float(img_width)
                
                if fit_mode == 'fit':
                    if (height/width) > aspect:
                        new_width = width
                        new_height = width * aspect
                    else:
                        new_height = height
                        new_width = height / aspect
                    x = (width - new_width) / 2
                    y = (height - new_height) / 2
                    c.drawImage(ImageReader(img), x, y, width=new_width, height=new_height)
                elif fit_mode == 'fill':
                    c.drawImage(ImageReader(img), 0, 0, width=width, height=height)
                else:
                    c.drawImage(ImageReader(img), 0, 0)
                
                c.showPage()
            except Exception as e:
                return jsonify({'error': f'Error processing image: {str(e)}'}), 400
        
        c.save()
        return send_file(output_path, as_attachment=True, download_name='converted_images.pdf')
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 2. MERGE PDFs
@app.route('/api/merge-pdf', methods=['POST'])
@track_usage('merge_pdf')
def merge_pdfs():
    try:
        files = request.files.getlist('pdfs')
        if len(files) < 2:
            return jsonify({'error': 'Please upload at least 2 PDFs'}), 400
        
        writer = PdfWriter()
        
        for file in files:
            try:
                reader = PdfReader(file.stream)
                for page in reader.pages:
                    writer.add_page(page)
            except Exception as e:
                return jsonify({'error': f'Error reading PDF {file.filename}: {str(e)}'}), 400
        
        output_path = os.path.join(UPLOAD_FOLDER, f"merged_{int(time.time())}.pdf")
        with open(output_path, 'wb') as f:
            writer.write(f)
        
        return send_file(output_path, as_attachment=True, download_name='merged.pdf')
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 3. SPLIT PDF
@app.route('/api/split-pdf', methods=['POST'])
@track_usage('split_pdf')
def split_pdf():
    try:
        file = request.files.get('pdf')
        if not file:
            return jsonify({'error': 'No PDF uploaded'}), 400
        
        split_type = request.form.get('split_type', 'all')
        pages_str = request.form.get('pages', '')
        
        reader = PdfReader(file.stream)
        total_pages = len(reader.pages)
        
        memory_file = io.BytesIO()
        
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            if split_type == 'all':
                for i, page in enumerate(reader.pages):
                    writer = PdfWriter()
                    writer.add_page(page)
                    page_io = io.BytesIO()
                    writer.write(page_io)
                    page_io.seek(0)
                    zf.writestr(f'page_{i+1}.pdf', page_io.getvalue())
            else:
                pages_to_extract = []
                try:
                    for part in pages_str.split(','):
                        part = part.strip()
                        if '-' in part:
                            start, end = map(int, part.split('-'))
                            pages_to_extract.extend(range(start-1, end))
                        else:
                            pages_to_extract.append(int(part)-1)
                except:
                    return jsonify({'error': 'Invalid page range format. Use: 1-3,5,7'}), 400
                
                for i in pages_to_extract:
                    if 0 <= i < total_pages:
                        writer = PdfWriter()
                        writer.add_page(reader.pages[i])
                        page_io = io.BytesIO()
                        writer.write(page_io)
                        page_io.seek(0)
                        zf.writestr(f'page_{i+1}.pdf', page_io.getvalue())
        
        memory_file.seek(0)
        return send_file(memory_file, as_attachment=True, download_name='split_pages.zip', mimetype='application/zip')
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 4. ROTATE PDF
@app.route('/api/rotate-pdf', methods=['POST'])
@track_usage('rotate_pdf')
def rotate_pdf():
    try:
        file = request.files.get('pdf')
        if not file:
            return jsonify({'error': 'No PDF uploaded'}), 400
        
        angle = int(request.form.get('angle', 90))
        
        reader = PdfReader(file.stream)
        writer = PdfWriter()
        
        for page in reader.pages:
            page.rotate(angle)
            writer.add_page(page)
        
        output_path = os.path.join(UPLOAD_FOLDER, f"rotated_{int(time.time())}.pdf")
        with open(output_path, 'wb') as f:
            writer.write(f)
        
        return send_file(output_path, as_attachment=True, download_name='rotated.pdf')
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 5. COMPRESS PDF
@app.route('/api/compress-pdf', methods=['POST'])
@track_usage('compress_pdf')
def compress_pdf():
    try:
        file = request.files.get('pdf')
        if not file:
            return jsonify({'error': 'No PDF uploaded'}), 400
        
        file_content = file.read()
        original_size = len(file_content)
        
        reader = PdfReader(io.BytesIO(file_content))
        writer = PdfWriter()
        
        for page in reader.pages:
            writer.add_page(page)
        
        if reader.metadata:
            writer.add_metadata(reader.metadata)
        
        output_path = os.path.join(UPLOAD_FOLDER, f"compressed_{int(time.time())}.pdf")
        with open(output_path, 'wb') as f:
            writer.write(f)
        
        compressed_size = os.path.getsize(output_path)
        savings = round((1 - compressed_size/original_size) * 100, 1) if original_size > 0 else 0
        
        response = send_file(output_path, as_attachment=True, download_name='compressed.pdf')
        response.headers['X-Original-Size'] = str(original_size)
        response.headers['X-Compressed-Size'] = str(compressed_size)
        response.headers['X-Savings'] = str(savings)
        
        return response
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 6. PDF INFO
@app.route('/api/pdf-info', methods=['POST'])
@track_usage('pdf_info')
def pdf_info():
    try:
        file = request.files.get('pdf')
        if not file:
            return jsonify({'error': 'No PDF uploaded'}), 400
        
        file_content = file.read()
        file_size_kb = round(len(file_content) / 1024, 2)
        
        reader = PdfReader(io.BytesIO(file_content))
        
        info = {
            'pages': len(reader.pages),
            'size_kb': file_size_kb,
            'encrypted': reader.is_encrypted,
            'author': '',
            'title': '',
            'creator': '',
            'width_mm': None,
            'height_mm': None
        }
        
        if reader.metadata:
            meta = reader.metadata
            info['author'] = str(meta.get('/Author', '')) if meta.get('/Author') else ''
            info['title'] = str(meta.get('/Title', '')) if meta.get('/Title') else ''
            info['creator'] = str(meta.get('/Creator', '')) if meta.get('/Creator') else ''
        
        if reader.pages:
            try:
                box = reader.pages[0].mediabox
                info['width_mm'] = round(float(box.width) * 0.352777, 1)
                info['height_mm'] = round(float(box.height) * 0.352777, 1)
            except:
                pass
        
        return jsonify(info)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 7. PASSWORD PDF
@app.route('/api/password-pdf', methods=['POST'])
@track_usage('password_pdf')
def password_pdf():
    try:
        file = request.files.get('pdf')
        if not file:
            return jsonify({'error': 'No PDF uploaded'}), 400
        
        password = request.form.get('password', '')
        action = request.form.get('action', 'encrypt')
        
        if not password:
            return jsonify({'error': 'Password is required'}), 400
        
        reader = PdfReader(file.stream)
        writer = PdfWriter()
        
        for page in reader.pages:
            writer.add_page(page)
        
        output_path = os.path.join(UPLOAD_FOLDER, f"password_{int(time.time())}.pdf")
        
        if action == 'encrypt':
            writer.encrypt(password)
            with open(output_path, 'wb') as f:
                writer.write(f)
            return send_file(output_path, as_attachment=True, download_name='protected.pdf')
        else:
            if reader.is_encrypted:
                try:
                    reader.decrypt(password)
                    writer = PdfWriter()
                    for page in reader.pages:
                        writer.add_page(page)
                except:
                    return jsonify({'error': 'Incorrect password'}), 400
            
            with open(output_path, 'wb') as f:
                writer.write(f)
            return send_file(output_path, as_attachment=True, download_name='decrypted.pdf')
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 8. PDF TO IMAGES
@app.route('/api/pdf-to-images', methods=['POST'])
@track_usage('pdf_to_images')
def pdf_to_images():
    try:
        file = request.files.get('pdf')
        if not file:
            return jsonify({'error': 'No PDF uploaded'}), 400
        
        fmt = request.form.get('format', 'JPEG')
        dpi = int(request.form.get('dpi', 150))
        pages_str = request.form.get('pages', 'all')
        
        pdf_bytes = file.read()
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        
        memory_file = io.BytesIO()
        
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            if pages_str == 'all':
                page_nums = range(len(doc))
            else:
                try:
                    if '-' in pages_str:
                        start, end = map(int, pages_str.split('-'))
                        page_nums = range(start-1, end)
                    else:
                        page_nums = [int(p.strip())-1 for p in pages_str.split(',')]
                except:
                    return jsonify({'error': 'Invalid page range. Use: 1-5 or 1,3,5'}), 400
            
            for i in page_nums:
                if 0 <= i < len(doc):
                    page = doc.load_page(i)
                    mat = fitz.Matrix(dpi/72, dpi/72)
                    pix = page.get_pixmap(matrix=mat)
                    
                    if fmt == 'PNG':
                        img_data = pix.tobytes("png")
                        ext = 'png'
                    else:
                        img_data = pix.tobytes("jpeg")
                        ext = 'jpg'
                    
                    zf.writestr(f'page_{i+1}.{ext}', img_data)
        
        doc.close()
        memory_file.seek(0)
        return send_file(memory_file, as_attachment=True, download_name='pdf_images.zip', mimetype='application/zip')
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 10000))
    app.run(host='0.0.0.0', port=port)
